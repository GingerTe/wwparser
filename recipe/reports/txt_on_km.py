# encoding: utf-8
from collections import OrderedDict
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from sqlalchemy.orm import sessionmaker

from engine import engine
from model import Data

Session = sessionmaker()
Session.configure(bind=engine)
session = Session()

wb = Workbook()
sheet = wb.active
txt = []
MAX_KM = 100
km = list(range(MAX_KM + 1))
sheet.append(['Текст/км'] + [str(x) for x in km])

res = {}

for data in session.query(Data).filter(Data.date > date(2018, 9, 19)).all():
    if data.txt.txt not in res:
        res[data.txt.txt] = OrderedDict.fromkeys(range(120), 0)
    res[data.txt.txt][data.km] += 1

for key in sorted(res):
    sheet.append([key] + list(res[key].values()))
    # cell = sheet.cell(row=txt.index(txt) + 1, column=km.index(data.km) + 1)
    # cell.value = (cell.value or 0) + 1

session.close()
sheet.column_dimensions['A'].width = 72
index = 0
for index, row in enumerate(sheet.rows):
    if not index:
        continue
    for cell in row:
        if not cell.value:
            cell.value = ''
        cell.alignment = Alignment(wrap_text=True)

tab = Table(displayName="Table1", ref="A1:{}{}".format(get_column_letter(MAX_KM), index + 1))

# Add a default style with striped rows and banded columns
style = TableStyleInfo(name="TableStyleLight16", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=True)
tab.tableStyleInfo = style
sheet.add_table(tab)

wb.save('km.xlsx')
